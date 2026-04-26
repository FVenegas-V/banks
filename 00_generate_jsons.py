#!/usr/bin/env python3
"""
PASO 0 — Extracción y chunking semántico desde PDFs.

Entradas:
  Datos_prueba/**/*.pdf

Salidas:
  logs/documents.json       — metadata por documento (1 entrada por PDF)
  logs/chunks.json          — chunks con texto crudo y posición
  logs/extraction_report.json — estadísticas y warnings

Diseño:
  - Chunking por párrafos con target ~600 chars, overlap ~100 chars.
  - Preserva page_start/page_end por chunk para citación auditable.
  - Detecta PDFs protegidos/vacíos y los reporta sin abortar.
  - Normaliza texto: une guiones de cierre de línea, colapsa whitespace,
    remueve numeración de página huérfana.
  - Tipo de documento e institución se derivan de la ruta + filename.
    El enriquecimiento semántico ocurre en 01_enrich_metadata.py.

Uso:
  python3 00_generate_jsons.py
"""
from __future__ import annotations

import json
import re
import sys
from dataclasses import asdict
from pathlib import Path
from typing import Optional

from models import Chunk, Document

# ---------------------------------------------------------------------------
# Configuración del chunker
# ---------------------------------------------------------------------------

TARGET_CHUNK_CHARS = 600     # tamaño deseado
MIN_CHUNK_CHARS = 200        # descarta/fusiona chunks más chicos
MAX_CHUNK_CHARS = 1200       # divide si un chunk crece más que esto
OVERLAP_CHARS = 100          # solapamiento entre chunks consecutivos

DATA_DIR = Path("Datos_prueba")
OUTPUT_DIR = Path("logs")
DOCUMENTS_JSON = OUTPUT_DIR / "documents.json"
CHUNKS_JSON = OUTPUT_DIR / "chunks.json"
REPORT_JSON = OUTPUT_DIR / "extraction_report.json"

# Patrones que indican PDF protegido / no extraíble
PROTECTED_MARKERS = (
    "Microsoft Azure Information Protection",
    "This is a protected document",
    "You can view it using a supported PDF reader",
)

# Patrón para título-de-sección-en-mayúsculas (líneas cortas en caps).
# Ej: "DECISIÓN", "ANTECEDENTES", "DISCUSIÓN Y ACUERDOS".
SECTION_TITLE_RE = re.compile(
    r"^([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑ\s]{2,60})\s*$"
)

# Fechas ISO-ish en nombre de archivo o texto
DATE_PATTERNS = [
    # 2022-07-13, 13-07-2022
    re.compile(r"(\d{4})[-_/](\d{1,2})[-_/](\d{1,2})"),
    re.compile(r"(\d{1,2})[-_/](\d{1,2})[-_/](\d{4})"),
    # 13 de julio de 2022
    re.compile(
        r"(\d{1,2})\s+de\s+(enero|febrero|marzo|abril|mayo|junio|"
        r"julio|agosto|septiembre|octubre|noviembre|diciembre)"
        r"\s+de\s+(\d{4})",
        re.IGNORECASE,
    ),
]

MONTH_ES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "septiembre": 9, "octubre": 10,
    "noviembre": 11, "diciembre": 12,
}


# ---------------------------------------------------------------------------
# Extracción
# ---------------------------------------------------------------------------

def extract_pages(pdf_path: Path) -> tuple[list[str], list[str]]:
    """Devuelve (páginas_de_texto, warnings)."""
    try:
        from pypdf import PdfReader
    except ImportError:  # compat: intentamos PyPDF2 como fallback
        try:
            from PyPDF2 import PdfReader  # type: ignore
        except ImportError:
            print("❌ Falta pypdf. Instala: pip install pypdf", file=sys.stderr)
            raise

    warnings: list[str] = []
    pages: list[str] = []
    try:
        with open(pdf_path, "rb") as f:
            reader = PdfReader(f)
            if reader.is_encrypted:
                try:
                    reader.decrypt("")
                except Exception:
                    warnings.append("pdf_encrypted")
                    return pages, warnings

            for page in reader.pages:
                try:
                    pages.append(page.extract_text() or "")
                except Exception as e:
                    warnings.append(f"page_extract_error: {e}")
                    pages.append("")
    except Exception as e:
        warnings.append(f"pdf_open_error: {e}")
        return pages, warnings

    # Detección de PDFs protegidos por rights management (texto placeholder)
    full_text = "\n".join(pages)
    if any(marker in full_text for marker in PROTECTED_MARKERS):
        warnings.append("pdf_rights_protected")

    # PDFs sin texto extraíble (probablemente escaneados, requieren OCR)
    if len(full_text.strip()) < 500 and not warnings:
        warnings.append("low_text_content_maybe_scanned")

    return pages, warnings


def normalize_page_text(text: str) -> str:
    """Limpia texto de una página."""
    if not text:
        return ""

    # Unir palabras cortadas por guión al final de línea: "inflació-\nn" -> "inflación"
    text = re.sub(r"-\n(\w)", r"\1", text)

    # Colapsar saltos intra-línea dentro de párrafos (no dobles)
    # Preservamos \n\n como separador de párrafo.
    text = re.sub(r"(?<!\n)\n(?!\n)", " ", text)

    # Colapsar espacios múltiples
    text = re.sub(r"[ \t]+", " ", text)

    # Eliminar líneas que son solo un número (page numbers huérfanos)
    text = re.sub(r"^\s*\d{1,3}\s*$", "", text, flags=re.MULTILINE)

    # Eliminar watermarks DRM de JPMorgan y similares: {[{<hash>}]}
    text = re.sub(r"\{\[\{[^}]*\}\]\}", "", text)

    return text.strip()


# ---------------------------------------------------------------------------
# Detección de tipo, institución, fecha
# ---------------------------------------------------------------------------

def detect_doc_type(rel_path: str) -> str:
    p = rel_path.lower()
    if "comunicado" in p:
        return "COMUNICADO"
    if "minuta" in p:
        return "MINUTA"
    if "/fed/" in p or p.startswith("fed/") or "reunion_fed" in p:
        return "FED_STATEMENT"
    if "monitor_pm" in p or "monitor pm" in p:
        return "MONITOR_PM"
    if "research" in p or "jpm" in p:
        return "REPORTE_RESEARCH"
    return "REPORTE_RESEARCH"  # default conservador


def detect_institution(rel_path: str) -> str:
    p = rel_path.lower()
    if "jpm" in p or "jpmorgan" in p:
        return "jpmorgan"
    if "/fed/" in p or p.startswith("fed/") or "fed" in p:
        return "federal_reserve"
    if "comunicado" in p or "minuta" in p or "monitor_pm" in p or "monitor pm" in p:
        return "banco_central_chile"
    return "unknown"


def detect_date(filename: str, first_page_text: str) -> Optional[str]:
    """Busca fecha en filename primero, luego en primeras 500 chars del doc."""
    haystack = filename + " " + first_page_text[:500]

    for pat in DATE_PATTERNS:
        m = pat.search(haystack)
        if not m:
            continue
        groups = m.groups()
        try:
            if len(groups) == 3 and groups[1] in MONTH_ES:
                day, month_name, year = groups
                return f"{int(year):04d}-{MONTH_ES[month_name.lower()]:02d}-{int(day):02d}"
            if len(groups[0]) == 4:
                y, m_, d = groups
                return f"{int(y):04d}-{int(m_):02d}-{int(d):02d}"
            d, m_, y = groups
            return f"{int(y):04d}-{int(m_):02d}-{int(d):02d}"
        except (ValueError, KeyError):
            continue

    # Fallback: solo año
    m = re.search(r"\b(19\d{2}|20\d{2})\b", haystack)
    return m.group(1) if m else None


def slugify_document_id(filename: str) -> str:
    base = Path(filename).stem.lower()
    return re.sub(r"[^a-z0-9]+", "_", base).strip("_")


# ---------------------------------------------------------------------------
# Chunker semántico
# ---------------------------------------------------------------------------

def split_paragraphs(text: str) -> list[str]:
    return [p.strip() for p in re.split(r"\n\s*\n+", text) if p.strip()]


def split_sentences(paragraph: str) -> list[str]:
    sentences = re.split(r"(?<=[.!?])\s+(?=[A-ZÁÉÍÓÚÑ0-9])", paragraph)
    return [s.strip() for s in sentences if s.strip()]


def detect_section_title(paragraph: str) -> Optional[str]:
    """Una línea corta en mayúsculas es probable título de sección."""
    if "\n" in paragraph:
        return None
    if len(paragraph) > 80:
        return None
    m = SECTION_TITLE_RE.match(paragraph)
    return m.group(1).strip() if m else None


def chunk_pages(pages: list[str]) -> list[dict]:
    """
    Chunking jerárquico:
      1. Itera página por página para preservar page_start/page_end.
      2. Divide cada página en párrafos.
      3. Acumula párrafos hasta llegar a TARGET_CHUNK_CHARS.
      4. Si un párrafo excede MAX_CHUNK_CHARS, lo divide por oraciones.
      5. Aplica overlap entre chunks consecutivos.

    Retorna lista de dicts: {text, page_start, page_end, section_title_raw}.
    """
    raw: list[dict] = []  # unidades pre-chunk con página asociada
    current_section: Optional[str] = None

    for page_num, page_text in enumerate(pages, start=1):
        cleaned = normalize_page_text(page_text)
        if not cleaned:
            continue

        for para in split_paragraphs(cleaned):
            title = detect_section_title(para)
            if title:
                current_section = title
                continue  # el título no es un chunk por sí mismo

            # Párrafo largo -> dividir en oraciones
            if len(para) > MAX_CHUNK_CHARS:
                for sent in split_sentences(para):
                    raw.append({"text": sent, "page": page_num, "section": current_section})
            else:
                raw.append({"text": para, "page": page_num, "section": current_section})

    # Agrupar raw units en chunks target
    chunks: list[dict] = []
    buffer: list[str] = []
    buf_pages: list[int] = []
    buf_section: Optional[str] = None
    buf_len = 0

    def flush():
        if not buffer:
            return
        text = " ".join(buffer).strip()
        if not text:
            return
        chunks.append({
            "text": text,
            "page_start": min(buf_pages),
            "page_end": max(buf_pages),
            "section_title_raw": buf_section,
        })

    for unit in raw:
        unit_len = len(unit["text"])

        # Si el buffer está casi lleno y este unit lo empujaría al límite, flush
        if buf_len + unit_len > TARGET_CHUNK_CHARS and buf_len >= MIN_CHUNK_CHARS:
            flush()
            # Overlap: conservar cola del buffer anterior
            if OVERLAP_CHARS > 0 and buffer:
                tail = " ".join(buffer)[-OVERLAP_CHARS:]
                buffer = [tail]
                buf_len = len(tail)
                buf_pages = [buf_pages[-1]]
            else:
                buffer = []
                buf_pages = []
                buf_len = 0
            buf_section = unit["section"] or buf_section

        buffer.append(unit["text"])
        buf_pages.append(unit["page"])
        buf_len += unit_len + 1
        if buf_section is None:
            buf_section = unit["section"]

    flush()

    # Fusión de chunks huérfanos: si el último es muy chico, fundirlo con el anterior
    if len(chunks) >= 2 and len(chunks[-1]["text"]) < MIN_CHUNK_CHARS:
        last = chunks.pop()
        prev = chunks[-1]
        prev["text"] = prev["text"] + " " + last["text"]
        prev["page_end"] = max(prev["page_end"], last["page_end"])

    return chunks


# ---------------------------------------------------------------------------
# Extracción Excel
# ---------------------------------------------------------------------------

def _find_header_row(rows: list[tuple]) -> tuple[int, list[str]]:
    """Devuelve (índice_fila_header, lista_de_headers). Ignora filas vacías al inicio."""
    for idx, row in enumerate(rows):
        non_empty = [c for c in row if c is not None and str(c).strip()]
        if len(non_empty) >= 2:
            headers = [str(c).strip() if c is not None else "" for c in row]
            return idx, headers
    return -1, []


def _cell_to_str(value) -> str:
    """Convierte un valor de celda a string limpio."""
    if value is None:
        return ""
    import datetime as dt
    if isinstance(value, (dt.datetime, dt.date)):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def extract_cell_chunks(excel_path: Path) -> tuple[list[dict], list[str]]:
    """
    Extrae chunks a nivel de celda desde un Excel con estructura:
      filas = fechas, columnas = categorías de texto narrativo.

    Cada celda con contenido genera un chunk independiente:
      text = "[YYYY-MM-DD - Nombre Columna]\n<texto de la celda>"
      section_title_raw = nombre de la columna
      page_start / page_end = número de hoja (1-based)

    Devuelve (lista_de_raw_chunks, warnings).
    raw_chunk keys: text, page_start, page_end, section_title_raw, fecha_iso
    """
    warnings: list[str] = []
    raw_chunks: list[dict] = []

    try:
        import openpyxl
    except ImportError:
        warnings.append("missing_openpyxl: pip install openpyxl")
        return raw_chunks, warnings

    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    except Exception as e:
        warnings.append(f"excel_open_error: {e}")
        return raw_chunks, warnings

    for sheet_idx, sheet_name in enumerate(wb.sheetnames, start=1):
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        header_idx, headers = _find_header_row(rows)
        if header_idx == -1:
            warnings.append(f"sheet_no_header: {sheet_name}")
            continue

        # Detectar columna de fecha (primera columna no vacía cuyo header
        # contiene "fecha" o sea la columna 0/1 con datetimes)
        fecha_col: Optional[int] = None
        for ci, h in enumerate(headers):
            if "fecha" in h.lower():
                fecha_col = ci
                break
        if fecha_col is None:
            fecha_col = 0  # fallback: primera columna

        # Columnas de contenido textual (todo excepto fecha)
        content_cols = [
            (ci, h) for ci, h in enumerate(headers)
            if ci != fecha_col and h
        ]

        for row in rows[header_idx + 1:]:
            if not any(c is not None for c in row):
                continue  # fila vacía

            fecha_val = row[fecha_col] if fecha_col < len(row) else None
            fecha_str = _cell_to_str(fecha_val)
            if not fecha_str:
                continue  # fila sin fecha, no indexable

            for ci, col_name in content_cols:
                if ci >= len(row):
                    continue
                cell_text = _cell_to_str(row[ci])
                if not cell_text or len(cell_text) < 10:
                    continue  # celda vacía o trivial

                text = f"[{fecha_str} - {col_name}]\n{cell_text}"
                raw_chunks.append({
                    "text": text,
                    "page_start": sheet_idx,
                    "page_end": sheet_idx,
                    "section_title_raw": col_name,
                    "fecha_iso": fecha_str,
                })

    wb.close()
    return raw_chunks, warnings


def process_excel(excel_path: Path) -> tuple[Optional[Document], list[Chunk]]:
    """
    Extrae chunks desde un .xlsx con filas=fechas y columnas=categorías de texto.
    Cada celda con contenido genera un chunk independiente.
    """
    rel_path = str(excel_path.relative_to(DATA_DIR))
    filename = excel_path.name
    doc_id = slugify_document_id(filename)

    raw_chunks, warnings = extract_cell_chunks(excel_path)

    if not raw_chunks:
        doc = Document(
            document_id=doc_id,
            filename=filename,
            filepath=rel_path,
            doc_type_category=detect_doc_type(rel_path),
            institution=detect_institution(rel_path),
            document_date=detect_date(filename, ""),
            total_pages=0,
            total_chunks=0,
            char_count=0,
            extraction_warnings=warnings or ["excel_no_content"],
        )
        return doc, []

    chunks: list[Chunk] = []
    for i, rc in enumerate(raw_chunks):
        text = rc["text"]
        chunks.append(Chunk(
            chunk_id=f"{doc_id}_{i:04d}",
            document_id=doc_id,
            text=text,
            char_count=len(text),
            token_estimate=max(1, len(text) // 4),
            page_start=rc["page_start"],
            page_end=rc["page_end"],
            position_in_doc=i,
            section_title_raw=rc["section_title_raw"],
        ))

    # Fecha del documento = primera fecha encontrada en los chunks
    first_fecha = raw_chunks[0]["fecha_iso"] if raw_chunks else ""
    total_chars = sum(c.char_count for c in chunks)
    total_sheets = max(rc["page_end"] for rc in raw_chunks) if raw_chunks else 0

    doc = Document(
        document_id=doc_id,
        filename=filename,
        filepath=rel_path,
        doc_type_category=detect_doc_type(rel_path),
        institution=detect_institution(rel_path),
        document_date=first_fecha or detect_date(filename, ""),
        total_pages=total_sheets,
        total_chunks=len(chunks),
        char_count=total_chars,
        extraction_warnings=warnings,
    )
    return doc, chunks


# ---------------------------------------------------------------------------
# Pipeline principal
# ---------------------------------------------------------------------------

def process_pdf(pdf_path: Path) -> tuple[Optional[Document], list[Chunk]]:
    rel_path = str(pdf_path.relative_to(DATA_DIR))
    filename = pdf_path.name
    doc_id = slugify_document_id(filename)

    pages, warnings = extract_pages(pdf_path)

    if not pages or "pdf_rights_protected" in warnings or "pdf_encrypted" in warnings:
        doc = Document(
            document_id=doc_id,
            filename=filename,
            filepath=rel_path,
            doc_type_category=detect_doc_type(rel_path),
            institution=detect_institution(rel_path),
            document_date=detect_date(filename, ""),
            total_pages=len(pages),
            total_chunks=0,
            char_count=0,
            extraction_warnings=warnings,
        )
        return doc, []

    first_page = pages[0] if pages else ""
    doc_type = detect_doc_type(rel_path)
    institution = detect_institution(rel_path)
    doc_date = detect_date(filename, first_page)

    raw_chunks = chunk_pages(pages)

    chunks: list[Chunk] = []
    for i, rc in enumerate(raw_chunks):
        text = rc["text"]
        chunks.append(Chunk(
            chunk_id=f"{doc_id}_{i:04d}",
            document_id=doc_id,
            text=text,
            char_count=len(text),
            token_estimate=max(1, len(text) // 4),
            page_start=rc["page_start"],
            page_end=rc["page_end"],
            position_in_doc=i,
            section_title_raw=rc["section_title_raw"],
        ))

    total_chars = sum(c.char_count for c in chunks)
    doc = Document(
        document_id=doc_id,
        filename=filename,
        filepath=rel_path,
        doc_type_category=doc_type,
        institution=institution,
        document_date=doc_date,
        total_pages=len(pages),
        total_chunks=len(chunks),
        char_count=total_chars,
        extraction_warnings=warnings,
    )
    return doc, chunks


def _build_chunk_char_stats(all_chunks: list[Chunk]) -> dict:
    char_counts = sorted(c.char_count for c in all_chunks)
    n = len(char_counts) or 1
    return {
        "min": char_counts[0] if char_counts else 0,
        "p25": char_counts[n // 4] if char_counts else 0,
        "p50_median": char_counts[n // 2] if char_counts else 0,
        "p75": char_counts[(3 * n) // 4] if char_counts else 0,
        "p90": char_counts[(9 * n) // 10] if char_counts else 0,
        "max": char_counts[-1] if char_counts else 0,
        "mean": sum(char_counts) // n if char_counts else 0,
    }


def _build_extraction_report(
    pdf_files: list[Path],
    excel_files: list[Path],
    all_docs: list[Document],
    all_chunks: list[Chunk],
    failed: list[str],
) -> dict:
    return {
        "pdf_count": len(pdf_files),
        "excel_count": len(excel_files),
        "documents_processed": len(all_docs),
        "documents_failed": failed,
        "documents_with_warnings": [
            {"document_id": doc.document_id, "warnings": doc.extraction_warnings}
            for doc in all_docs if doc.extraction_warnings
        ],
        "chunks_total": len(all_chunks),
        "chunk_char_stats": _build_chunk_char_stats(all_chunks),
        "chunking_config": {
            "target_chunk_chars": TARGET_CHUNK_CHARS,
            "min_chunk_chars": MIN_CHUNK_CHARS,
            "max_chunk_chars": MAX_CHUNK_CHARS,
            "overlap_chars": OVERLAP_CHARS,
        },
    }


def _save_outputs(
    all_docs: list[Document],
    all_chunks: list[Chunk],
    stats: dict,
) -> None:
    with open(DOCUMENTS_JSON, "w", encoding="utf-8") as f:
        json.dump([asdict(doc) for doc in all_docs], f, ensure_ascii=False, indent=2)
    with open(CHUNKS_JSON, "w", encoding="utf-8") as f:
        json.dump([asdict(chunk) for chunk in all_chunks], f, ensure_ascii=False, indent=2)
    with open(REPORT_JSON, "w", encoding="utf-8") as f:
        json.dump(stats, f, ensure_ascii=False, indent=2)


EXCEL_EXTENSIONS = {".xlsx", ".xls"}


def main() -> int:
    OUTPUT_DIR.mkdir(exist_ok=True)

    pdf_files = sorted(DATA_DIR.rglob("*.pdf"))
    excel_files = sorted(
        p for ext in EXCEL_EXTENSIONS for p in DATA_DIR.rglob(f"*{ext}")
    )
    all_files = pdf_files + excel_files

    if not all_files:
        print(f"❌ No hay PDFs ni Excel en {DATA_DIR}/", file=sys.stderr)
        return 1

    print(f"[00] Procesando {len(pdf_files)} PDFs y {len(excel_files)} Excel desde {DATA_DIR}/")

    all_docs: list[Document] = []
    all_chunks: list[Chunk] = []
    failed: list[str] = []

    for file_path in all_files:
        rel = file_path.relative_to(DATA_DIR)
        if file_path.suffix.lower() in EXCEL_EXTENSIONS:
            doc, chunks = process_excel(file_path)
        else:
            doc, chunks = process_pdf(file_path)
        if doc is None:
            print(f"  ✗ {rel} — no procesable")
            failed.append(str(rel))
            continue
        all_docs.append(doc)
        all_chunks.extend(chunks)
        warn_suffix = f" [warnings: {','.join(doc.extraction_warnings)}]" if doc.extraction_warnings else ""
        print(f"  ✓ {rel} — {len(chunks)} chunks, {doc.total_pages} páginas{warn_suffix}")

    stats = _build_extraction_report(pdf_files, excel_files, all_docs, all_chunks, failed)
    _save_outputs(all_docs, all_chunks, stats)

    char_stats = stats["chunk_char_stats"]
    print(f"\n[00] ✓ {DOCUMENTS_JSON} ({len(all_docs)} documentos)")
    print(f"[00] ✓ {CHUNKS_JSON} ({len(all_chunks)} chunks)")
    print(f"[00] ✓ {REPORT_JSON}")
    print(f"[00] chunk size: min={char_stats['min']} "
          f"p50={char_stats['p50_median']} "
          f"p90={char_stats['p90']} "
          f"max={char_stats['max']}")
    if failed:
        print(f"[00] ⚠ {len(failed)} PDFs fallaron: {failed}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
